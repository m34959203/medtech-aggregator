"""Загрузчик ОФИЦИАЛЬНОГО справочника услуг (MedArchive, от организаторов).

Формат `Справочник услуг.xlsx`: ID · Специальность · Code · Name_ru · TarificatrCode
(1286 услуг, ~573 специальности). Синонимов нет → нормализация по коду + нечётко.

Каждая строка Name_ru → запись ServiceCatalog с tarificator_code + specialty.
Один TarificatrCode может встречаться у нескольких услуг (3D/4D УЗИ → один код) —
это нормально, маппинг по коду затем доуточняется fuzzy по имени.

Идемпотентно: повторный запуск не плодит дубли (upsert по (tarificator_code, name)).
"""
from __future__ import annotations

import io

from sqlalchemy.orm import Session

from ..models import ServiceCatalog
from .archive_extractor import norm_code

# Специальность → грубая категория витрины (для совместимости с MedPrice-категориями).
_CATEGORY_BY_SPEC = [
    (("лаборатор", "анализ", "гематолог", "биохим", "микробиолог", "пцр", "иммун", "гистолог", "цитолог"), "Анализы"),
    (("узи", "ультразвук"), "УЗИ"),
    (("мрт", "магнитно-резонанс"), "МРТ/КТ"),
    (("кт", "компьютерная томограф"), "МРТ/КТ"),
    (("рентген", "флюорограф", "маммограф"), "Рентген"),
    (("стоматолог", "ортодонт", "зубн"), "Стоматология"),
]


# Имена официального справочника, которые должны влиться в УЖЕ существующий
# SPECS-канон, а НЕ создавать отдельный дубль (долг #41/#43). Ключ/значение — lower.
# Напр. справочник зовёт «Глюкоза (кровь)», а витринный канон — «Глюкоза (в крови)».
_CANON_ALIASES = {
    "глюкоза (кровь)": "глюкоза (в крови)",
    "глюкоза (кровь) экспресс": "глюкоза (в крови)",
}


def _category_for(specialty: str, name: str) -> str:
    blob = f"{specialty} {name}".lower()
    for keys, cat in _CATEGORY_BY_SPEC:
        if any(k in blob for k in keys):
            return cat
    # приём/консультация специалиста
    if any(k in blob for k in ("прием", "приём", "консультац")):
        return "Приём врача"
    return "Прочее"


def _read_rows(path_or_bytes) -> list[dict]:
    import openpyxl

    if isinstance(path_or_bytes, (bytes, bytearray)):
        src = io.BytesIO(path_or_bytes)
    else:
        src = path_or_bytes
    # read_only=False сознательно — иначе на некоторых .xlsx ws.cell даёт None
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    i_spec = col("специальн", "specialty")
    i_name = col("name_ru", "наименован", "название", "услуг")
    i_code = col("tarificatr", "тарификат", "код")
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        name = vals[i_name] if i_name is not None and i_name < len(vals) else None
        if name is None or not str(name).strip():
            continue
        spec = str(vals[i_spec]).strip() if i_spec is not None and vals[i_spec] is not None else ""
        code = norm_code(vals[i_code]) if i_code is not None and i_code < len(vals) else None
        rows.append({"name": str(name).strip(), "specialty": spec, "code": code})
    return rows


def load_official_catalog(db: Session, path_or_bytes) -> dict:
    """Загружает/обновляет справочник в ServiceCatalog. → статистика."""
    rows = _read_rows(path_or_bytes)
    # существующие — индекс по (canonical_name) для идемпотентности
    existing = {s.canonical_name.strip().lower(): s for s in db.query(ServiceCatalog).all()}
    created = updated = 0
    for row in rows:
        key = row["name"].lower()
        # алиас: вливаем в существующий канон вместо создания дубля
        alias = _CANON_ALIASES.get(key)
        if alias and alias in existing:
            svc = existing[alias]
            if row["name"].lower() not in {s.lower() for s in (svc.synonyms or [])}:
                svc.synonyms = list(svc.synonyms or []) + [row["name"]]
            if row["code"] and not svc.tarificator_code:
                svc.tarificator_code = row["code"]
            updated += 1
            continue
        svc = existing.get(key)
        category = _category_for(row["specialty"], row["name"])
        if svc:
            changed = False
            if row["code"] and svc.tarificator_code != row["code"]:
                svc.tarificator_code = row["code"]
                changed = True
            if row["specialty"] and svc.specialty != row["specialty"]:
                svc.specialty = row["specialty"]
                changed = True
            if not svc.category or svc.category == "Прочее":
                svc.category = category
                changed = True
            updated += int(changed)
        else:
            svc = ServiceCatalog(
                canonical_name=row["name"],
                category=category,
                synonyms=[],
                tarificator_code=row["code"],
                specialty=row["specialty"],
            )
            db.add(svc)
            existing[key] = svc
            created += 1
    db.flush()
    with_code = sum(1 for r in rows if r["code"])
    return {"rows": len(rows), "created": created, "updated": updated, "with_code": with_code}
